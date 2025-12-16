import os
import io
import zipfile
import pandas as pd
import pyodbc
import logging
import uuid
import numpy as np  # [중요] name 'np' is not defined 오류 해결을 위한 import 구문
from flask import Flask, render_template, request, jsonify, send_file, session
from datetime import timedelta
from openpyxl.utils import get_column_letter  # 엑셀 컬럼 너비 조절을 위해 추가
from openpyxl.styles import Font              # 엑셀 폰트 스타일링을 위해 추가
import openpyxl  # 파일 상단에 import 되어 있는지 확인

# --- 기본 설정 ---
logging.basicConfig(level=logging.INFO)
app = Flask(__name__)
# app.secret_key = os.urandom(24)  # 이 라인을 주석 처리하거나 삭제하고
# 아래와 같이 고정된 문자열로 변경합니다.
app.secret_key = 'dev-secret-key-any-string-is-ok'
app.permanent_session_lifetime = timedelta(minutes=30)
TEMP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_data')
if not os.path.exists(TEMP_DIR):
    os.makedirs(TEMP_DIR)

# --- 웹 페이지 및 API 기능 정의 ---

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/get_companies', methods=['POST'])
def get_companies():
    # ... (이전과 코드 동일) ...
    db_config = request.json.get('db_config')
    if not db_config:
        return jsonify({"error": "DB 정보가 없습니다."}), 400
    try:
        conn_str = (f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={db_config["server"]};DATABASE={db_config["database"]};UID={db_config["uid"]};PWD={db_config["password"]};TrustServerCertificate=yes;')
        with pyodbc.connect(conn_str, timeout=10) as cnxn:
            company_df = pd.read_sql("SELECT co_cd, co_nm FROM sco ORDER BY co_nm", cnxn)
            companies = company_df.to_dict(orient='records')
            return jsonify({"companies": companies})
    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        if 'S0002' in sqlstate:
             return jsonify({"error": f"연결은 성공했으나 'sco' 테이블을 찾을 수 없습니다."}), 500
        return jsonify({"error": f"DB 연결 실패: {ex}"}), 500
    except Exception as e:
        return jsonify({"error": f"알 수 없는 오류 발생: {e}"}), 500

# [수정] 메인 데이터 조회 기능
@app.route('/api/fetch', methods=['POST'])
def fetch_data():
    """선택된 쿼리에 맞춰 동적으로 파라미터를 생성하여 데이터를 조회합니다."""
    try:
        data = request.json
        db_config = data.get('db_config')
        
        # --- [핵심 변경점] 선택된 메뉴 이름에 따라 sql과 params를 각각 설정 ---
        query_name = data.get('query_name')
        co_cd = data.get('co_cd')
        co_nm = data.get('co_nm')
        start_date = data.get('start_date')
        end_date = data.get('end_date')

        sql = None
        params = None

        if query_name == "거래처등록":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT ISNULL(CO_CD, '') AS CO_CD, ISNULL(TR_FG, '') AS TR_FG, ISNULL(TR_CD, '') AS TR_CD, "
                "ISNULL(TR_NM, '') AS TR_NM, ISNULL(ATTR_NM, '') AS ATTR_NM, ISNULL(CEO_NM, '') AS CEO_NM, "
                "ISNULL(REG_NB, '') AS REG_NB, ISNULL(PPL_NB, '') AS PPL_NB, ISNULL(FOR_YN, '') AS FOR_YN, "
                "ISNULL(BUSINESS, '') AS BUSINESS, ISNULL(JONGMOK, '') AS JONGMOK, ISNULL(ZIP, '') AS ZIP, "
                "ISNULL(DIV_ADDR1, '') AS DIV_ADDR1, ISNULL(ADDR2, '') AS ADDR2, ISNULL(TEL, '') AS TEL, "
                "ISNULL(FAX, '') AS FAX, ISNULL(HOMEPAGE, '') AS HOMEPAGE, ISNULL(EMAIL, '') AS EMAIL, "
                "ISNULL(LIQ_RS, '') AS LIQ_RS, ISNULL(PJT_CD, '') AS PJT_CD, ISNULL(TRGRP_CD, '') AS TRGRP_CD, "
                "ISNULL(TRRAT_FG, '') AS TRRAT_FG, ISNULL(CLTTR_CD, '') AS CLTTR_CD, ISNULL(LOCAL_CD, '') AS LOCAL_CD, "
                "ISNULL(LINK_CD, '') AS LINK_CD, ISNULL(JEONJA_YN, '') AS JEONJA_YN, ISNULL(REPTR_CD, '') AS REPTR_CD, "
                "ISNULL(APPR_NB, '') AS APPR_NB, ISNULL(LIQ_FG, '') AS LIQ_FG, ISNULL(REMARK_DC, '') AS REMARK_DC, "
                "ISNULL(INTER_DT, '') AS INTER_DT, ISNULL(DUE_DT, '') AS DUE_DT, ISNULL(USE_YN, '') AS USE_YN, "
                "ISNULL(TRSO_FG, '') AS TRSO_FG, ISNULL(PAYCON_DC, '') AS PAYCON_DC, ISNULL(DOUDATE1_FG, '') AS DOUDATE1_FG, "
                "ISNULL(DOUDATE1_DD, '') AS DOUDATE1_DD, ISNULL(EMP_CD, '') AS EMP_CD, ISNULL(DEPT_CD, '') AS DEPT_CD, "
                "'' AS TITLE_DC, '' AS JOB_DC, '' AS TEL_DC, '' AS EXT_DC, '' AS HP_DC, '' AS EMAIL_DC, "
                "'' AS RMK_DC, ISNULL(RECZIP, '') AS RECZIP, ISNULL(REC_ADDR1, '') AS REC_ADDR1, ISNULL(REC_ADDR2, '') AS REC_ADDR2, "
                "'' AS REC_TEL, '' AS REC_HP, '' AS REC_FAX, '' AS REC_MAIL, '' AS EMPGRP_CD, '' AS TRCHARGE_EMP, "
                "'' AS TRCHARGE_EMAIL, '' AS TRCHARGE_DEPT, '' AS TRCHARGE_TITLE, '' AS TRCHARGE_JOP, ISNULL(TRCHARGE_TEL, '') AS TRCHARGE_TEL, "
                "ISNULL(TRCHARGE_EXT, '') AS TRCHARGE_EXT, '' AS TRCHARGE_HP, ISNULL(JIRO_CD, '') AS JIRO_CD, ISNULL(BA_NB, '') AS BA_NB, "
                "ISNULL(DEPOSITOR, '') AS DEPOSITOR, ISNULL(ACCT1_FG, '') AS ACCT1_FG, ISNULL(TAX1_FG, '') AS TAX1_FG, ISNULL(SETTLE_PAY, '') AS SETTLE_PAY, "
                "ISNULL(SETTLE_CD, '') AS SETTLE_CD, ISNULL(ACCT2_FG, '') AS ACCT2_FG, ISNULL(TAX2_FG, '') AS TAX2_FG, ISNULL(SETTLE_RCP, '') AS SETTLE_RCP "
                "FROM STRADE WHERE CO_CD = ? "
            )
            params = [co_cd]

        elif query_name == "사원정보":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT ROW_NUMBER() OVER (ORDER BY EMP_CD) AS 'NO', EMP_CD AS 사번, KOR_NM AS 이름, "
                "'' AS 로그인ID, '' AS 메일ID, '' AS 로그인비밀번호, '' AS 결재비밀번호, "
                "CASE WHEN GENDER_FG = 'M' THEN '남자' ELSE '여자' END AS '성별', '한국어' AS 사용언어, "
                "GRUP_DT AS '최초입사일', '' AS 기본라이선스, '' AS 메일, '' AS 원챔버탐색기팩, "
                "'' AS ONEAI라이선스, HCLS_CD AS 직급, HRSP_CD AS 직책, ENRL_FG AS 재직구분, "
                "EMPL_FG AS 고용구분, HOPR_CD AS 직무, JOIN_DT AS 입사일, '사용' AS 근태사용, "
                "REFR_YN AS 내외국인여부, RSRG_NO AS 주민등록번호, FORG_NO AS 외국인등록번호, "
                "PYSP_FG AS 급여형태, HTYP_CD AS 직종 FROM SEMP WHERE CO_CD = ? AND EMPL_FG = '001'"
            )
            params = [co_cd]

        elif query_name == "사원정보2":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT '' AS [순번], EMP_CD AS [사번], KOR_NM AS [이름(한국어)], ENLS_NM AS [이름(영어)], "
                "'' AS [이름(일본어)], '' AS [이름(중국어)], '' AS [로그인ID], '' AS [메일ID], "
                "'' AS [로그인 비밀번호], '' AS [결재 비밀번호], "
                "CASE WHEN GENDER_FG = 'M' THEN '남자' ELSE '여자' END AS [성별], '한국어' AS [사용언어], "
                "EMAL_ADD AS [개인메일], '' AS [휴대전화], TEL AS [전화번호], "
                "'' AS [비상연락망(HP) 관계], EMGC_TEL AS [비상연락망(HP)], RR_ZIP AS [우편번호], "
                "RSRG_ADD AS [기본주소], RSRD_ADD AS [상세주소], '' AS [우편번호(영문)], "
                "'' AS [기본주소(영문)], '' AS [상세주소(영문)], JOIN_DT AS [최초입사일], "
                "RTR_DT AS [최종퇴사일], '미사용' AS [계정 사용], '미사용' AS [모바일 사용], "
                "'미사용' AS [메신저 사용], BCCN_NO AS [출입카드번호], '' AS [SNS], "
                "'' AS [커리어], '비라이선스' AS [기본 라이선스], '미사용' AS [메일], "
                "'미사용' AS [ONE-AI 라이선스], '미사용' AS [원챔버탐색기팩], '1000' AS [회사코드], "
                "'1000' AS [사업장코드], DEPT_CD AS [부서코드], HCLS_CD AS [직급코드], "
                "HRSP_CD AS [직책코드], ENRL_FG AS [재직구분코드], EMPL_FG AS [고용구분코드], "
                "HOPR_CD AS [직무코드], JOIN_DT AS [입사일], RPMC_DT AS [퇴사일], "
                "'' AS [회사전화번호], '' AS [회사팩스번호], '' AS [내선단축번호], "
                "'' AS [회사우편번호], '' AS [회사기본주소], '' AS [회사상세주소], "
                "'' AS [회사우편번호(영문)], '' AS [회사기본주소(영문)], '' AS [회사상세주소(영문)], "
                "'미표시' AS [조직도], '미표시' AS [대화/쪽지 조직도], '미사용' AS [근태사용], "
                "REFR_YN AS [내외국인여부], RSRG_NO AS [주민등록번호], FORG_NO AS [외국인등록번호], "
                "'' AS [외국인단일세율적용여부], '' AS [생년월일], '' AS [양/음력구분], "
                "'' AS [급여이메일], '' AS [본적주소], '' AS [본적주소상세], "
                "'' AS [본적우편번호], '' AS [수습적용여부], '' AS [수습만료일], "
                "'' AS [수습급여지급률], '' AS [근속기간포함여부(수습)], '' AS [휴직시작일], "
                "'' AS [휴직종료일], '' AS [휴직복직일], '' AS [휴직사유], "
                "'' AS [휴직급여지급율], '' AS [퇴직기간포함여부(휴직)], '' AS [중도정산일], "
                "'' AS [계약종료일], '' AS [퇴직사유], '' AS [계정유형], '' AS [호봉], "
                "PYSP_FG AS [급여형태], HTYP_CD AS [직종], '' AS [근무조], '' AS [프로젝트], "
                "'' AS [하위사업], '' AS [분류코드], '' AS [사용자정의1], '' AS [사용자정의2], "
                "'' AS [사용자정의3], '' AS [사용자정의4], '' AS [사용자정의5], "
                "'' AS [사용자정의6], '' AS [국적코드(관리용)], '' AS [국적코드(신고용)], "
                "'' AS [거주구분], '' AS [거주지국코드], '' AS [세대주여/부], "
                "'' AS [최종학력], '' AS [국외소득유무], '' AS [건강보험번호], "
                "'' AS [건강보험료자동계산여부], '' AS [건강보험보수월액], '' AS [건강보험금액], "
                "'' AS [건강보험여부], '' AS [노인장기요양보험료], '' AS [장기요양보험료 경감대상자여부], "
                "'' AS [연금유형], '' AS [국민연금자동계산여부], '' AS [국민연금보수월액], "
                "'' AS [국민연금금액], '' AS [사학연금자동계산여부], '' AS [사학연금보수월액], "
                "'' AS [사학연금], '' AS [고용보험여부], '' AS [고용보험 자동계산여부], "
                "'' AS [고용보험보수월액], '' AS [고용보험료], '' AS [고용보험대표이사여부], "
                "'' AS [노조가입여부], '' AS [(급여)이체은행], '' AS [(급여)계좌번호], "
                "'' AS [예금주], '' AS [(기타)이체은행], '' AS [(기타)계좌번호], "
                "'' AS [예금주2], '' AS [장애인구분], '' AS [배우자공제 여부], "
                "'' AS [20세이하 인원수], '' AS [60세이상 인원수], "
                "'' AS [장애인기본공제인원수], '' AS [기타수급자], '' AS [다자녀추가공제], "
                "'' AS [부녀자공제], '' AS [학자금공제자 여부], '' AS [학자금상환 통지액], "
                "'' AS [생산직 총급여], '' AS [퇴직연금구분], '' AS [퇴직연금은행코드], "
                "'' AS [퇴직연금계좌번호], '' AS [퇴직연금가입일], '' AS [퇴직연금해지일], "
                "'' AS [퇴직연금 비고], '' AS [두루누리사회보험 신청여부], "
                "'' AS [두루누리사회보험적용율], '' AS [자금입력방식], "
                "'' AS [경비이체은행코드], '' AS [경비이체계좌번호], "
                "'' AS [경비이체예금주], '' AS [종교관련종사자여부], "
                "'' AS [근무지], '' AS [학자금공제기간시작], "
                "'' AS [학자금공제기간종료], '' AS [감면대상], "
                "'' AS [감면유형코드], '' AS [감면기간(시작연월)], "
                "'' AS [감면기간(종료연월)], '' AS [담당업무], "
                "'' AS [담당업무(영문)], '' AS [담당업무(중국어)], "
                "'' AS [담당업무(일본어)], '' AS [외국법인소속파견자여부], "
                "'' AS [결혼여부], '' AS [결혼일], '' AS [기타근무유형], "
                "'' AS [기타근무시간] FROM SEMP WHERE CO_CD = ? AND EMPL_FG = '001'"
            )
            params = [co_cd]

        elif query_name == "조직정보":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT D.CO_CD, C.CO_NM, H.CTRL_NM, '' AS 상세구분, D.CTD_CD, D.CTD_NM, "
                "'' AS 영어, '' AS 중국어, '' AS 일본어, 'Y', '' AS 정렬 "
                "FROM SCTRL_D D "
                "INNER JOIN SCO C ON D.CO_CD = C.CO_CD "
                "LEFT OUTER JOIN SCTRL H ON H.CO_CD = D.CO_CD AND H.CTRL_CD = D.CTRL_CD AND H.MODULE_CD = 'H' "
                "WHERE D.CO_CD = ? AND D.CTRL_CD IN ('G3','G4','G5','G6') AND H.MODULE_CD = 'H' "
            )
            params = [co_cd]

        elif query_name == "상용직정보":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT '' AS [순번], '' AS [로그인ID], '' AS [메일ID], '' AS [로그인 비밀번호], "
                "KOR_NM AS [이름(한국어)], ENLS_NM AS [이름(영어)], '' AS [이름(일본어)], "
                "'' AS [이름(중국어)], CASE WHEN GENDER_FG = 'M' THEN '남자' ELSE '여자' END AS [성별], "
                "'' AS [휴대전화], TEL AS [전화번호], EMGC_TEL AS [비상연락망(HP)], "
                "'' AS [관계], RR_ZIP AS [우편번호], RSRG_ADD AS [기본주소], RSRD_ADD AS [상세주소], "
                "'한국어' AS [사용언어], EMAL_ADD AS [개인메일], '' AS [개인메일도메인], "
                "EMAL_ADD AS [급여메일], '' AS [급여메일도메인], '비라이선스' AS [기본 라이선스], "
                "'N' AS [메일], 'N' AS [원챔버 탐색기팩], 'N' AS [OneAI], JOIN_DT AS [최초입사일], "
                "RTR_DT AS [최종퇴사일], '' AS [회사코드], DEPT_CD AS [부서코드], EMP_CD AS [사번], "
                "HCLS_CD AS [직급코드], HRSP_CD AS [직책코드], ENRL_FG AS [재직구분코드], "
                "EMPL_FG AS [고용구분코드], HOPR_CD AS [직무코드], '' AS [담당업무(한국어)], "
                "'' AS [담당업무(영어)], '' AS [담당업무(중국어)], '' AS [담당업무(일본어)], "
                "JOIN_DT AS [입사일], RPMC_DT AS [퇴사일], '미사용' AS [근태사용], "
                "TEL AS [전화번호(조직)], '' AS [내선번호(조직)], '' AS [팩스번호(조직)], "
                "RR_ZIP AS [우편번호(조직)], RSRG_ADD AS [기본주소(조직)], RSRD_ADD AS [상세주소(조직)], "
                "'표시' AS [조직도], '표시' AS [대화/쪽지조직도] FROM SEMP WHERE CO_CD = ?  AND EMPL_FG = '001'"
            )
            params = [co_cd]

        elif query_name == "부서정보":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT '' AS 회사코드, '' AS 회사명, '' AS 사업장코드, '' AS 사업장명, "
                "'-' AS 상위부서코드, 1 AS Depth, dept_cd AS 부서코드, dept_nm AS 부서명, "
                "'부서' AS 유형 FROM SDEPT WHERE CO_CD = ? "
            )
            params = [co_cd]

        elif query_name == "품목등록":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT ITEM_CD, ITEM_NM, ITEM_DC, ACCT_FG, ODR_FG, ITEMGRP_CD, UNIT_DC, UNITMANG_DC, UNITCHNG_NB, USE_YN, "
                "LEAD_DT, DESIGN_NB, LOT_FG, SETITEM_FG, QC_FG, '0' REQ_FG, TRMAIN_CD, STANDARD_UM, REAL_UM, SALE_UM, "
                "PURCH_UM, SALPLN_CD, PURPLN_CD, MTLPLN_CD, MFGPLN_CD, EMP_CD, L_CD, M_CD, S_CD, LOT_QT, FOQ_QT, "
                "TRSUB_CD, TRSUB2_CD, SAFESTOCK_QT, COMP_DC, BARCODE_DC, WIDTH_QT, LENGTH_QT, HEIGHT_QT, DENSTY_QT, "
                "WIDTH_UM, LENGTH_UM, HEIGHT_UM, VOLUME_QT, WEIGHT_QT, AREA_QT, VOLUME_UM, WEIGHT_UM, AREA_UM, "
                "DAYCAPA_NB, LNSP_UM, HS_NB, MASS_CD, MASS_CONFNB, PACK_PO_QT, PACK_PO_RT, PACK_SO_QT, PACK_SO_RT, "
                "STANDARD_TIME,'' aaa,LIQ_YN, REMARK_DC, AL_REMARK_DC, LIQ_FG, STANDARD_TIME_UNIT_DC, SALEVAT_UM, PURCHVAT_UM, "
                "'' LIQUSE_FG FROM SITEM WHERE CO_CD = ?"
            )
            params = [co_cd]

        elif query_name == "품목군등록":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT ITEMGRP_CD, ITEMGRP_NM, USE_YN, ITEMGRP_DC FROM SITEMGRP WHERE CO_CD = ?"
            )
            params = [co_cd]
        
        elif query_name == "BOM등록":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT ITEMPARENT_CD	,BOM_SQ SORT_SQ	,ITEMCHILD_CD	,JUST_QT	,LOSS_RT	,OUT_FG	,USE_YN	,REAL_QT	,ODR_FG	,START_DT	,END_DT	,REMARK_DC	FROM SBOM_WF WHERE CO_CD = ? and itemparent_cd in (SELECT item_cd FROM SITEM WHERE co_cd = ? and use_yn = '1')"
            )
            params = [co_cd,co_cd]

        elif query_name == "관리내역등록":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT H.CTRL_CD,D.MGM_CD,D.MGM_NM,D.REMARK_DC,D.USE_YN FROM LCTRL_MGM H INNER JOIN  LCTRL_MGM_D D ON H.CO_CD = D.CO_CD AND H.CTRL_CD = D.CTRL_CD WHERE D.CO_CD = ?"
            )
            params = [co_cd]

        elif query_name == "창고":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT DIV_CD,BASELOC_FG,BASELOC_CD,BASELOC_NM,BASELOC_DC,USE_YN FROM SBASELOC WHERE CO_CD = ?"
            )
            params = [co_cd]

        elif query_name == "공정":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT BASELOC_CD,LOC_CD,LOC_NM,TR_CD,BAD_YN,AVABSTOCK_YN,USE_YN,LOC_DC FROM SLOC WHERE CO_CD = ?"
            )
            params = [co_cd]

        elif query_name == "프로젝트등록":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT CO_CD,PJT_CD,PROG_FG,PJT_NM,PJT_NMK,ATPJT_NM,ATPJT_NMK,TR_CD,PJTGRP_CD,FR_DT,TO_DT,START_DT,PJT_WORKTY,PJTRMK_DC,PJTRMK_DCK,COST_FG,PJT_TY,DEPT_CD,EMP_CD,'' PRIME_TR_CD,ORD_AM,GOYONG_NM,RSRG_NO,HJS,HCLS,DTL_DC,GOAL_DC FROM SPJT  WHERE CO_CD = ?"
            )
            params = [co_cd]
        elif query_name == "납품처등록":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT CUST_CD,	SHIP_CD,	SHIP_NM,	ZIP_CD,	ADD_DC,	SHIP_FG,	CARGO_AM,	TREMP_NM,	PHONE_NB,	FAX_NB,	CP_NB,	EMAIL_DC,	USE_YN FROM LSHIP_BILL  WHERE CO_CD = ?"
            )
            params = [co_cd]

        elif query_name == "물류담당자등록":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT PLN_CD,PLN_NM,EMP_CD,PLN_TEL,PLN_FAX,PLN_CP,PLNS_CD,FROM_DT,TO_DT,USE_YN FROM LPLNNERCD  WHERE CO_CD = ?"
            )
            params = [co_cd]
        elif query_name == "고객별물류담당자등록":
            if not co_cd:
                return jsonify({"error": "회사를 선택해야 합니다."}), 400
            sql = (
                "SELECT TR_CD,PLN_CD,PURPLN_CD,OUTPLN_CD,AREA_CD,AREA_GRP,TRADE_GRP,SHIP_CD,UM_FG,POUM_FG   FROM LTRADEMGM  WHERE CO_CD = ?"
            )
            params = [co_cd]
        elif query_name == "기초재고":
            sql = (
                "SELECT LI.DIV_CD, LEFT(?, 4) + '0101' AS ADJUST_DT, LI.WH_CD, LI.LC_CD,'기초재고' AS REMARK_DC,'' AS TR_CD,'' AS PLN_CD, LI.ITEM_CD,SUM(ISNULL(LI.IOPEN_QT, 0)) AS IOPEN_QT, MAX(ISNULL(T.OPEN_UM, 0)) AS OPEN_UM, "
                "CASE WHEN ROW_NUMBER() OVER (PARTITION BY LI.DIV_CD, LI.ITEM_CD ORDER BY LI.LOT_NB) = 1 THEN MAX(ISNULL(T.OPEN_AM, 0)) ELSE 0 END AS OPEN_AM, "
                "'' AS MGMT_CD, '' AS PJT_CD, ISNULL(LI.LOT_NB, '') AS LOT_NB "
                "FROM LX_LINVTORY LI "
                "LEFT OUTER JOIN LINV_TAV T ON LI.CO_CD = T.CO_CD AND LI.DIV_CD = T.DIV_CD AND LI.P_YR + '01' = T.SMM AND LI.ITEM_CD = T.ITEM_CD "
                "WHERE LI.CO_CD = ? AND LI.P_YR = LEFT(?, 4) "
                "GROUP BY LI.DIV_CD, LI.WH_CD, LI.LC_CD, LI.ITEM_CD, LI.LOT_NB "
                "HAVING SUM(LI.IOPEN_QT) <> 0 ORDER BY LI.WH_CD,LI.LC_CD,LI.ITEM_CD, LI.LOT_NB"
            )
            params = [
                start_date,
                co_cd,
                start_date
            ]

        elif query_name == "주문정보":
            if not all([co_cd, start_date, end_date]):
                return jsonify({"error": "회사,시작일자,종료일자 모두 선택해야 합니다."}), 400
            sql = (
                "SELECT H.SO_FG, H.SO_DT, H.TR_CD, 'KRW' AS EXCH_CD, ISNULL(H.EXCH_RT,1) EXCH_RT, H.VAT_FG, H.UMVAT_FG, "
                "ISNULL(H.SHIP_CD,'') SHIP_CD, ISNULL(H.PLN_CD,'') PLN_CD, ISNULL(H.LC_NB,'') LC_NB, "
                "ISNULL(D.SO_NB,'') MGM_NM, ISNULL(H.REMARK_DC,'') REMARK_DC, D.ITEM_CD, D.DUE_DT, D.SHIPREQ_DT, "
                "ISNULL(D.SO_QT,0) SO_QT, ISNULL(D.UM_FG,'') UM_FG, ISNULL(D.SO_UM,0) SO_UM, D.VAT_UM, D.SOG_AM, "
                "D.SOV_AM, D.SOH_AM, D.EXCH_UM, D.EXCH_AM, '0' REQ_FG, ISNULL(D.QC_FG,'0') QC_FG, D.MGMT_CD, D.PJT_CD, "
                "ISNULL(LEFT(D.REMARK_DC,57),'') + '|'+CONVERT(NVARCHAR(3),D.SO_SQ) REMARKD_DC, D.EST_NB, "
                "CASE WHEN D.EST_SQ= 0 THEN NULL ELSE D.EST_SQ END EST_SQ, D.ITEMSET_CD "
                "FROM LSO H "
                "LEFT OUTER JOIN LSO_D D ON H.CO_CD = D.CO_CD AND H.SO_NB = D.SO_NB "
                "WHERE H.SO_FG = '0' AND H.CO_CD = ? AND H.SO_DT >=  ? AND H.SO_DT <= ? "
                "ORDER BY H.CO_CD, H.SO_NB, D.SO_SQ"
            )
            params = [co_cd, start_date, end_date]

        elif query_name == "입고처리":
            if not all([co_cd, start_date, end_date]):
                return jsonify({"error": "회사,시작일자,종료일자 모두 선택해야 합니다."}), 400
            sql = (
                "SELECT H.DIV_CD,H.PO_FG, H.RCV_DT, H.TR_CD, H.EXCH_CD, ISNULL(H.EXCH_RT,1) EXCH_RT, H.VAT_FG, H.UMVAT_FG, "
                "H.WH_CD, H.PLN_CD,H.RCV_NB REMARK_DC, D.ITEM_CD, D.PO_QT, D.RCV_QT, D.UM_FG, D.RCV_UM, D.VAT_UM, "
                "D.RCVG_AM, D.RCVV_AM, D.RCVH_AM, D.EXCH_UM, D.EXCH_AM, D.LC_CD, D.LOT_NB,H.MGMT_CD, D.PJT_CD, "
                "D.REMARK_DC REMARKD_DC, PO_NB, PO_SQ,  IBL_NB, IBL_SQ, '' REQ_NB, '' REQ_SQ, "
                "'' QC_NB, '' QC_SQ "
                "FROM LSTOCK H "
                "LEFT OUTER JOIN LSTOCK_D D ON H.CO_CD = D.CO_CD AND H.RCV_NB = D.RCV_NB "
                "WHERE H.CO_CD = ? AND H.RCV_DT >=  ? AND H.RCV_DT <= ? "
                "ORDER BY H.CO_CD, d.RCV_NB, D.RCV_SQ"
            )
            params = [co_cd, start_date, end_date]

        elif query_name == "발주등록":
            if not all([co_cd, start_date, end_date]):
                return jsonify({"error": "회사,시작일자,종료일자 모두 선택해야 합니다."}), 400
            sql = (
                "SELECT H.PO_DT, H.TR_CD, H.PO_FG, H.VAT_FG, H.EXCH_CD, D.ITEM_CD, D.DUE_DT, D.PO_QT, D.EXCH_UM, D.EXCH_AM, ISNULL(H.EXCH_RT,1) EXCH_RT, ISNULL(H.MGMT_CD,'') MGMT_CD, ISNULL(D.PJT_CD,'') PJT_CD, ISNULL(H.LC_NB,'') LC_NB, ISNULL(H.PLN_CD,'') PLN_CD, H.REMARK_DC, D.PO_UM, ISNULL(D.UM_FG,'') UM_FG, D.POG_AM, D.POGV_AM1, D.POGH_AM1, D.SHIPREQ_DT, '' REQ_FG, '' QC_FG, ISNULL(D.REMARK_DC,'') + '|' + CONVERT(NVARCHAR(3),D.PO_SQ) REMARKD_DC, '' SO_NB, '' SO_SQ, '' REQ_NB, '' MREQ_SQ, '' REQR_NB, '' REQR_SQ, TRNMTD_DC, PAYCON_DC, DELARA_DC,D.PO_NB MGM_NM, UMVAT_FG, VAT_UM " 
                "FROM LPO H INNER JOIN LPO_D D ON H.CO_CD = D.CO_CD AND H.PO_NB = D.PO_NB "
                "WHERE H.CO_CD = ? AND H.PO_DT >= ? AND H.PO_DT <= ? "
                "ORDER BY H.CO_CD, H.PO_NB, D.PO_SQ"
            )
            params = [co_cd, start_date, end_date]

        elif query_name == "출고처리":
            if not all([co_cd, start_date, end_date]):
                return jsonify({"error": "회사,시작일자,종료일자 모두 선택해야 합니다."}), 400
            sql = (
                "SELECT H.DIV_CD,H.SO_FG, H.ISU_DT, H.TR_CD, H.EXCH_CD, H.EXCH_RT, H.VAT_FG, H.UMVAT_FG, H.SHIP_CD, H.WH_CD, "
                "'' AS AREA_CD, H.SHIP_FG, H.PLN_CD, H.REMARK_DC, D.ITEM_CD, D.SO_QT, D.ISU_QT, D.UM_FG, D.ISU_UM, "
                "D.VAT_UM, D.ISUG_AM, D.ISUV_AM, D.ISUH_AM, D.EXCH_UM, D.EXCH_AM, D.LC_CD, D.LOT_NB, D.MGMT_CD, "
                "D.PJT_CD, D.REMARK_DC AS REMARKD_DC, D.SO_NB, D.SO_SQ, D.REQ_NB, D.REQ_SQ, D.QC_NB, D.QC_SQ "
                "FROM dbo.LDELIVER AS H "
                "INNER JOIN dbo.LDELIVER_D AS D ON H.CO_CD = D.CO_CD AND H.ISU_NB = D.ISU_NB "
                "WHERE H.CO_CD = ? AND H.ISU_DT >= ? AND H.ISU_DT <= ? "
                "ORDER BY H.CO_CD, D.ISU_NB, D.ISU_SQ"
            )
            params = [co_cd, start_date, end_date]
            
        elif query_name == "수금등록":
            if not all([co_cd, start_date, end_date]):
                return jsonify({"error": "회사,시작일자,종료일자 모두 선택해야 합니다."}), 400
            sql = (
               "SELECT H.RCP_FG RCPH_FG,H.RCP_DT,H.TR_CD,H.REF_DC,H.PLN_CD,H.REMARK_DC,D.RCP_FG,D.RCPMGTR_CD,D.RCPMG_DC,D.JATA_FG,D.NORMAL_AM,D.BEFORE_AM,"
                "D.BANK_CD,D.ISU_DT,D.DUE_DT,D.REMARK_DC REMARKD_DC,H.MGMT_CD,D.PJT_CD " 
                "FROM LRCP H INNER JOIN LRCP_D D ON H.CO_CD = D.CO_CD AND H.RCP_NB =D.RCP_NB "
                "WHERE H.RCP_FG = '0' AND H.CO_CD = ? AND H.rcp_DT BETWEEN ? AND ? "
                "ORDER BY H.RCP_NB"
            )
            params = [co_cd, start_date, end_date]
            
        elif query_name == "생산실적":
            if not all([co_cd, start_date, end_date]):
                return jsonify({"error": "회사,시작일자,종료일자 모두 선택해야 합니다."}), 400
            sql = (
                "SELECT DOC_DT, ITEM_CD, ITEM_QT, MOVEBASELOC_CD, MOVELOC_CD, LOT_NB, PJT_CD,'' MGMT_CD, "
                "PLN_CD,DOC_CD DOC_DC, BASELOC_CD, LOC_CD, EQUIP_CD "
                "FROM LORCV_H WHERE CO_CD = ? AND DOC_DT >=  ? AND DOC_DT <= ? "
                "UNION ALL "
                "SELECT DOC_DT,PITEM_CD,ITEM_QT,BASELOC_CD,LOC_CD,LOT_NB,PJT_CD,MGMT_CD,PLN_CD,DOC_CD REMARK_DC,WR_WH_CD,WR_LC_CD,EQUIP_CD FROM LPRODUCTION "
                "WHERE CO_CD = ? AND DOC_DT >= ? AND DOC_DT <= ?"
            )
            params = [co_cd, start_date, end_date,co_cd, start_date, end_date]

        elif query_name == "생산출고":
            if not all([co_cd, start_date, end_date]):
                return jsonify({"error": "회사,시작일자,종료일자 모두 선택해야 합니다."}), 400
            sql = (
                "SELECT U.USE_DT ISU_DT, U.ITEM_CD, U.USE_QT ISU_QT, '' FWH_CD, '' FLC_CD, '' PLN_CD, "
                "H.DOC_CD WR_CD, H.ITEM_CD ITEMPARENT_CD, '' LOT_NB, U.PJT_CD, U.MGMT_CD, U.REMARK_DC, "
                "'' WO_CD, '' MATL_SQ, '' OP_NB "
                "FROM LORCV_H H "
                "LEFT OUTER JOIN LMTL_USE U ON H.CO_CD = U.CO_CD AND H.DOC_CD = U.WR_CD "
                "WHERE H.CO_CD = ? AND U.USE_DT >= ? AND  U.USE_DT <= ? "
                "union all "
                "SELECT H.DOC_DT,D.CITEM_CD,D.USE_QT,D.BASELOC_CD,D.LOC_CD, '' PLN_CD,D.DOC_CD, H.PITEM_CD,D.LOT_NB,'' PJT_CD,'' MGMT_CD,D.REMARK_DC ,'' WO_CD, '' MATL_SQ, '' OP_NB "
                "FROM LPRODUCTION H INNER JOIN  LPRODUCTION_D D ON H.CO_CD = D.CO_CD AND H.DOC_CD = D.DOC_CD "
                "WHERE H.CO_CD = ? AND H.DOC_DT >= ? AND  H.DOC_DT <= ?"
            )
            params = [co_cd, start_date, end_date,co_cd, start_date, end_date]

        elif query_name == "재고조정":
            if not all([co_cd, start_date, end_date]):
                return jsonify({"error": "회사,시작일자,종료일자 모두 선택해야 합니다."}), 400
            sql = (
                "SELECT CASE WHEN H.ADJUST_FG='0' THEN '기초' WHEN H.ADJUST_FG='1' THEN '입고' WHEN H.ADJUST_FG='2' THEN '출고' END ADJUST_FG,H.DIV_CD, "
                "H.ADJUST_DT, H.WH_CD, H.LC_CD, H.REMARK_DC, H.TR_CD, H.PLN_CD, D.ITEM_CD, "
                "CASE WHEN H.ADJUST_FG ='0' THEN D.OPEN_QT WHEN H.ADJUST_FG = '1' THEN D.RCV_QT ELSE D.ISU_QT END ADJUST_QT, "
                "D.ADJUST_UM, D.ADJUST_AM, D.MGMT_CD, D.PJT_CD, D.LOT_NB, D.REMARK_DC REMARKD_DC "
                "FROM LADJUST H "
                "INNER JOIN LADJUST_D D ON H.CO_CD = D.CO_cD AND H.ADJUST_NB = D.ADJUST_NB "
                "WHERE H.CO_CD = ? AND H.ADJUST_DT >= ? AND H.ADJUST_DT <= ? AND ADJUST_FG IN ('0','1','2')"
            )
            params = [co_cd, start_date, end_date]
            
        elif query_name == "재고이동":
            if not all([co_cd, start_date, end_date]):
                return jsonify({"error": "회사,시작일자,종료일자 모두 선택해야 합니다."}), 400
            sql = (
                "SELECT H.MOVE_DT,H.PLN_CD,H.REMARK_DC,D.ITEM_CD,D.MOVE_QT,H.FWH_CD,H.FLC_CD," 
                "H.TWH_CD,H.TLC_CD,D.LOT_NB,D.PJT_CD,H.MGMT_CD,D.REMARK_DC REMARKD_DC "
                "FROM LSTKMOVE H INNER JOIN LSTKMOVE_D D ON H.CO_CD = D.CO_CD AND H.MOVE_NB = D.MOVE_NB "
                "WHERE H.CO_CD = ? AND H.MOVE_DT >= ? AND  H.MOVE_DT <= ?"
            )
            params = [co_cd, start_date, end_date]

        elif query_name == "회계초기이월":
            if not all([co_cd, start_date]):
                return jsonify({"error": "회사,시작일자,종료일자 모두 선택해야 합니다."}), 400
            sql = (
                "SELECT GISU,DIV_CD,ACCT_CD,ISU_SQ,DEPT_CD,EMP_CD,ACCT_AM,TR_CD,CT_DEPT,"
                "'' CASH_CD,PJT_CD,'' M_EMP_CD,'' EXCH_RT,'' EXCH_CD,'' EXCH_AM,'' BL_NO FROM APREV "
                "WHERE CO_CD = ? AND FILL_DT = LEFT(?,4)+'0101'"
            )
            params = [co_cd, start_date]

        elif query_name == "급여자료 추출":
            if not all([co_cd, start_date, end_date]):
                return jsonify({"error": "회사, 시작일, 종료일을 모두 선택해야 합니다."}), 400
            sql = "EXEC sp_GetDynamicPayrollPivot_Fixed ?, ?, ?"
            start_year = start_date[:4]
            end_year = end_date[:4]
            params = [co_cd, start_year, end_year]

        elif query_name == "자동전표처리":
            if not all([co_cd, start_date, end_date]):
                return jsonify({"error": "회사,시작일자,종료일자 모두 선택해야 합니다."}), 400
            sql = (
                "SELECT H.DIV_CD AS IN_DIV_CD, H.ISU_DT AS MENU_DT, H.ISU_SQ AS MENU_SQ, D.LN_SQ AS MENU_LN_SQ, "
                "H.ISU_DOC, H.DOCU_TY, D.DRCR_FG,LEFT(D.ACCT_CD,3) +'00'+ RIGHT(D.ACCT_CD,2) ACCT_CD, D.TR_CD, D.TR_NM, D.ACCT_AM, D.ATTR_CD, D.RMK_DC, "
                "T.DIV_CD VAT_DIV_CD, H.FILL_DT ISS_DT , D.CT_DEAL TAX_FG, T.NONSUB_TY, T.SUP_AM, T.DUMMY1, T.CASH_AM, "
                "T.BILL_AM, D.CT_NB,"
                "CASE WHEN D.CT_DEAL = '24' THEN D.CT_QT ELSE D.CT_AM END CT_AM, T.JEONJA_YN, T.ASSET_AM, T.ASSET_VAT, "
                "CASE WHEN AV2.ASSET_FG = '1' THEN AV2.ITEM_DC ELSE '' END AS ASSET_FG1, "
                "CASE WHEN AV2.ASSET_FG = '1' THEN ISNULL(AV2.MATT_CNT, 0) ELSE 0 END AS MATT_CNT1, "
                "CASE WHEN AV2.ASSET_FG = '1' THEN ISNULL(AV2.SUP_AM, 0) ELSE 0 END AS SUP_AM1, "
                "CASE WHEN AV2.ASSET_FG = '1' THEN ISNULL(AV2.VAT_AM, 0) ELSE 0 END AS VAT_AM1, "
                "CASE WHEN AV2.ASSET_FG = '2' THEN AV2.ITEM_DC ELSE '' END AS ASSET_FG2, "
                "CASE WHEN AV2.ASSET_FG = '2' THEN ISNULL(AV2.MATT_CNT, 0) ELSE 0 END AS MATT_CNT2, "
                "CASE WHEN AV2.ASSET_FG = '2' THEN ISNULL(AV2.SUP_AM, 0) ELSE 0 END AS SUP_AM2, "
                "CASE WHEN AV2.ASSET_FG = '2' THEN ISNULL(AV2.VAT_AM, 0) ELSE 0 END AS VAT_AM2, "
                "CASE WHEN AV2.ASSET_FG = '3' THEN AV2.ITEM_DC ELSE '' END AS ASSET_FG3, "
                "CASE WHEN AV2.ASSET_FG = '3' THEN ISNULL(AV2.MATT_CNT, 0) ELSE 0 END AS MATT_CNT3, "
                "CASE WHEN AV2.ASSET_FG = '3' THEN ISNULL(AV2.SUP_AM, 0) ELSE 0 END AS SUP_AM3, "
                "CASE WHEN AV2.ASSET_FG = '3' THEN ISNULL(AV2.VAT_AM, 0) ELSE 0 END AS VAT_AM3, "
                "CASE WHEN AV2.ASSET_FG = '4' THEN AV2.ITEM_DC ELSE '' END AS ASSET_FG4, "
                "CASE WHEN AV2.ASSET_FG = '4' THEN ISNULL(AV2.MATT_CNT, 0) ELSE 0 END AS MATT_CNT4, "
                "CASE WHEN AV2.ASSET_FG = '4' THEN ISNULL(AV2.SUP_AM, 0) ELSE 0 END AS SUP_AM4, "
                "CASE WHEN AV2.ASSET_FG = '4' THEN ISNULL(AV2.VAT_AM, 0) ELSE 0 END AS VAT_AM4, "
                "SB.BILL_NB, SB.ISS_DT ISSUE_DT, SB.DUE_DT, SB.BANK_CD FIN_CD, '' BILLCASH_CD, SB.ISU_NM, SB.ENDORS_NM, "
                "SB.BILL_FG1, SB.BILL_FG2, SB.DEAL_FG, "
                "CASE WHEN D.ACCT_CD IN ('13300','26300') THEN LEFT(D.CT_NB,3) + '00' + RIGHT(D.CT_NB,2) ELSE '' END LACCT_CD, CASE WHEN D.FR_DT = '00000000' THEN '' ELSE D.FR_DT END FR_DT, CASE WHEN D.TO_DT='00000000' THEN '' ELSE D.TO_DT END TO_DT, "
                "CASE WHEN D.ACCT_CD IN ('13300','26300') THEN D.CT_DEAL ELSE '' END CALC_TY, "
                "CASE WHEN D.DEPTCD_TY = 'C1' THEN D.CT_DEPT ELSE '' END CT_DEPT, "
                "CASE WHEN D.DEPTCD_TY = 'C2' THEN D.CT_DEPT ELSE '' END CASH_CD, "
                "CASE WHEN D.USER1_TY = 'L1' THEN D.CT_USER1 ELSE '' END CAR_CD, "
                "CASE WHEN D.DEPTCD_TY = 'C4' THEN D.CT_DEPT ELSE '' END FI_IN_CD, "
                "CASE WHEN D.DEPTCD_TY = 'C5' THEN D.CT_DEPT ELSE '' END SEC_CD, "
                "CASE WHEN D.DEPTCD_TY = 'C9' THEN D.CT_DEPT ELSE '' END LOCAL_CD, "
                "CASE WHEN D.PJTCD_TY = 'D1' THEN D.PJT_CD ELSE '' END PJT_CD, "
                "CASE WHEN D.PJTCD_TY = 'D3' THEN D.PJT_CD ELSE '' END WH_CD, "
                "CASE WHEN D.PJTCD_TY = 'D4' THEN D.PJT_CD ELSE '' END M_EMP_CD, "
                "CASE WHEN D.PJTCD_TY = 'D5' THEN D.PJT_CD ELSE '' END M_DIV_CD, "
                "CASE WHEN D.PJTCD_TY = 'D6' THEN D.PJT_CD ELSE '' END ITEMGRP_CD, "
                "CASE WHEN D.PJTCD_TY = 'D8' THEN D.PJT_CD ELSE '' END PROC_FG, "
                "CASE WHEN D.CTNB_TY = 'E2' THEN D.CT_NB ELSE '' END BL_NB, "
                "CASE WHEN D.CTNB_TY = 'E3' THEN D.CT_NB ELSE '' END LC_NB, "
                "CASE WHEN D.CTNB_TY = 'E4' THEN D.CT_NB ELSE '' END STOCK_NB, "
                "CASE WHEN D.CTNB_TY = 'E5' THEN LEFT(D.CT_NB,3) + '00' + RIGHT(D.CT_NB,2) ELSE ''  END SUB_ACCT, "
                "CASE WHEN D.CTNB_TY = 'E6' THEN D.CT_NB ELSE '' END ITEM_CD, "
                "CASE WHEN D.CTNB_TY = 'E7' THEN D.CT_NB ELSE '' END ASSET_CD, "
                "CASE WHEN D.CTNB_TY = 'E8' THEN D.CT_NB ELSE '' END LOAN_NB, "
                "CASE WHEN D.CTNB_TY = 'E9' THEN D.CT_NB ELSE '' END LEND_NB, "
                "CASE WHEN D.CTNB_TY = 'EA' THEN D.CT_NB ELSE '' END MA_NB, "
                "CASE WHEN D.FRDT_TY = 'F1' THEN D.FR_DT ELSE '' END OCCUR_DT, "
                "CASE WHEN D.FRDT_TY = 'F2' THEN D.FR_DT ELSE '' END START_DT, "
                "CASE WHEN D.FRDT_TY = 'F3' THEN D.FR_DT ELSE '' END REQ_DT, "
                "CASE WHEN D.TODT_TY = 'G1' THEN D.TO_DT ELSE '' END EXP_DT, "
                "CASE WHEN D.TODT_TY = 'G2' THEN D.TO_DT ELSE '' END END_DT, "
                "CASE WHEN D.TODT_TY = 'G3' THEN D.TO_DT ELSE '' END REPAY_DT, "
                "CASE WHEN D.QT_TY = 'H1' THEN D.CT_QT ELSE 0 END CT_QT, "
                "CASE WHEN D.QT_TY = 'H2' THEN D.CT_QT ELSE 0 END REPAY_QT, "
                "CASE WHEN D.QT_TY = 'H3' THEN D.CT_QT ELSE 0 END NDEDUCTION_AM, "
                "CASE WHEN D.QT_TY = 'H5' THEN D.CT_QT ELSE 0 END EXCH_RT, "
                "CASE WHEN D.QT_TY = 'H6' THEN D.CT_QT ELSE 0 END CARDRCP_AM, "
                "CASE WHEN D.AM_TY = 'I2' THEN D.CT_AM ELSE 0 END PARV_AM, "
                "CASE WHEN D.AM_TY = 'I3' THEN D.CT_AM ELSE 0 END STANDARD_AM, "
                "CASE WHEN D.AM_TY = 'I4' THEN D.CT_AM ELSE 0 END UNIT_AM, "
                "CASE WHEN D.AM_TY = 'I5' THEN D.CT_AM ELSE 0 END FORE_AM, "
                "CASE WHEN D.AM_TY = 'I6' THEN D.CT_AM ELSE 0 END PAY_AM, "
                "CASE WHEN D.RT_TY = 'J1' THEN D.CT_RT ELSE 0 END INTER_RT, "
                "CASE WHEN D.RT_TY = 'J2' THEN D.CT_RT ELSE 0 END DISCOUNT_RT, "
                "CASE WHEN D.RT_TY = 'J3' THEN D.CT_RT ELSE 0 END FEE_RT, "
                "CASE WHEN D.DEAL_TY = 'K1' THEN D.CT_DEAL ELSE '' END CT_DEAL, "
                "CASE WHEN D.DEAL_TY = 'K2' THEN D.CT_DEAL ELSE '' END EXCH_CD, "
                "CASE WHEN D.DEAL_TY = 'K3' THEN D.CT_DEAL ELSE '' END RECBILL_FG, "
                "CASE WHEN D.DEAL_TY = 'K4' THEN D.CT_DEAL ELSE '' END PAYBILL_FG, "
                "CASE WHEN D.DEAL_TY = 'K6' THEN D.CT_DEAL ELSE '' END OTHER_FG, "
                "CASE WHEN D.DEAL_TY = 'K7' THEN D.CT_DEAL ELSE '' END COLLECT_FG, "
                "CASE WHEN D.DEAL_TY = 'K8' THEN D.CT_DEAL ELSE '' END EVALUATION_FG, "
                "CASE WHEN D.DEAL_TY = 'K9' THEN D.CT_DEAL ELSE '' END CALC_FG, "
                "CT_USER1 USERL_TY1, '' USERL_TY2, '' USERL_TY3, '' USERL_TY4, '' USERL_TY5, '' USERL_TY6, '' USERL_TY7, '' USERL_TY8, '' USERL_TY9, "
                "CT_USER2 USERM_TY1, '' USERM_TY2, '' USERM_TY3, '' USERM_TY4, '' USERM_TY5, '' USERM_TY6, '' USERM_TY7, '' USERM_TY8, '' USERM_TY9, "
                "'' CEO_NM, '' REG_NB, '' BUSINESS, '' JONGMOK, '' DIV_ADDR1, '' ADDR2, '' TEL, '' EMAIL, '' EMPGRP_CD, '' TRCHARGE_EMP, "
                "'' TRCHARGE_EMAIL, '' TRCHARGE_DEPT, '' TRCHARGE_TEL, '' TRCHARGE_HP, '' ISS_NO, d.DIV_CD2, d.MGT_CD, d.BOTTOM_CD, d.BGT_CD, d.RMK_NB "
                "FROM ADOCUH H "
                "INNER JOIN ADOCUD D ON H.CO_CD = D.CO_CD AND H.ISU_DT = D.ISU_DT AND H.ISU_SQ = D.ISU_SQ "
                "LEFT OUTER JOIN ATAX T ON D.CO_CD = T.CO_CD AND D.ISU_DT = T.ISU_DT AND D.ISU_SQ = T.ISU_SQ AND D.LN_SQ = T.LN_SQ "
                "LEFT OUTER JOIN AVASSET2 AV2 ON D.CO_CD = AV2.CO_CD AND D.ISU_DT = AV2.ISU_DT AND D.ISU_SQ = AV2.ISU_SQ AND D.LN_SQ = AV2.LN_SQ "
                "LEFT OUTER JOIN SBILL SB ON D.CO_CD = SB.CO_CD AND D.ISU_DT = SB.ISU_DT AND D.ISU_SQ = SB.ISU_SQ AND D.LN_SQ = SB.LN_SQ "
                "WHERE D.CO_CD = ? AND CONVERT(DATE, H.ISU_DT, 112) >= ? AND H.ISU_DT <= ? " 
                "ORDER BY H.ISU_DT, H.ISU_SQ, D.LN_SQ"
            )
            params = [co_cd, start_date, end_date]

        else:
            return jsonify({"error": "유효하지 않은 쿼리 이름입니다."}), 400
        
        app.logger.info(f"실행할 쿼리: {sql}")
        app.logger.info(f"전달할 파라미터: {params}")
        # --- 변경점 끝 ---

        conn_str = (f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={db_config["server"]};DATABASE={db_config["database"]};UID={db_config["uid"]};PWD={db_config["password"]};TrustServerCertificate=yes;')
        with pyodbc.connect(conn_str, timeout=10) as cnxn:
            df = pd.read_sql(sql, cnxn, params=params)

        app.logger.info(f"데이터 조회 완료: {len(df)}개")
        
        df = df.replace({np.nan: None})

        data_id = str(uuid.uuid4())
        filepath = os.path.join(TEMP_DIR, f"{data_id}.parquet")
        df.to_parquet(filepath, engine='pyarrow')
        session['data_id'] = data_id
        session['query_name'] = query_name  #
        session['co_cd'] = co_cd 
        session['co_nm'] = co_nm 
        
        preview_df = df.head(100)
        
        return jsonify({
            "message": f"총 {len(df)}개 데이터 조회 완료. 상위 100개를 미리보기로 표시합니다.",
            "columns": list(preview_df.columns),
            "data": preview_df.to_dict(orient='records')
        })

    except Exception as e:
        app.logger.error(f"데이터 조회 중 오류: {e}", exc_info=True)
        return jsonify({"error": f"서버 오류: {str(e)}"}), 500

import openpyxl  # 파일 상단에 import 되어 있는지 확인
# export_excel 함수만 수정
@app.route('/api/export', methods=['POST'])
def export_excel():
    filepath = None
    try:
        data_id = session.get('data_id')
        query_name = session.get('query_name')
        co_cd = session.get('co_cd')
        co_nm = session.get('co_nm')

        if not all([data_id, query_name, co_cd]):
            return jsonify({"error": "데이터를 먼저 조회해야 합니다."}), 400

        filepath = os.path.join(TEMP_DIR, f"{data_id}.parquet")
        if not os.path.exists(filepath):
            return jsonify({"error": "서버에 데이터가 존재하지 않습니다. 다시 조회해주세요."}), 404

        df = pd.read_parquet(filepath, engine='pyarrow')
        if df.empty:
            return jsonify({"error": "변환할 데이터가 없습니다."}), 400

        data = request.get_json(silent=True)
        if data is None:
            data = {}
            
        split_rows = int(data.get('split_rows', 50000))

        download_filename = f'{co_cd}_{query_name}_data.zip'  # ZIP 파일명 생성
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED, False) as zf:
            for i in range(0, len(df), split_rows):
                chunk_df = df.iloc[i:i + split_rows]
                file_name = f"{co_cd}_{co_nm}_{query_name}_part_{i // split_rows + 1}.xlsx"

                # 급여자료 추출의 경우 템플릿을 사용하지 않고 새 엑셀 파일 생성
                if query_name == "급여자료 추출":
                    workbook = openpyxl.Workbook()
                    worksheet = workbook.active
                    worksheet.title = "급여자료"

                    # 헤더 작성
                    for col_idx, column_name in enumerate(chunk_df.columns, start=1):
                        worksheet.cell(row=1, column=col_idx, value=column_name)
                        worksheet.cell(row=1, column=col_idx).font = Font(bold=True)

                    # 데이터 작성
                    for row_idx, row_data in enumerate(chunk_df.itertuples(index=False), start=2):
                        for col_idx, cell_value in enumerate(row_data, start=1):
                            if pd.isna(cell_value):
                                cell_value = ""
                            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

                    # 열 너비 자동 조정
                    for col_idx, column_name in enumerate(chunk_df.columns, start=1):
                        max_length = max(
                            len(str(cell_value)) if cell_value else 0
                            for cell_value in [column_name] + chunk_df.iloc[:, col_idx - 1].astype(str).tolist()
                        )
                        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

                    # 엑셀 파일을 메모리에 저장
                    excel_buffer = io.BytesIO()
                    workbook.save(excel_buffer)
                    excel_buffer.seek(0)

                    zf.writestr(file_name, excel_buffer.getvalue())
                else:
                    # 기존 템플릿 기반 처리
                    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
                    TEMPLATE_FOLDER = os.path.join(BASE_DIR, 'excel_templates')

                    template_config = {
                        "거래처등록": {"file": "거래처등록_template.xlsx", "start_row": 4},
                        "사원정보": {"file": "사원정보_template.xlsx", "start_row": 8},
                        "부서정보": {"file": "부서정보_template.xlsx", "start_row": 22},
                        "사원정보2": {"file": "사원정보2_template.xlsx", "start_row": 8},
                        "조직정보": {"file": "조직정보_template.xlsx", "start_row": 8},
                        "주문정보": {"file": "주문정보_template.xlsx", "start_row": 4},
                        "생산실적": {"file": "생산입고_template.xlsx", "start_row": 4},
                        "생산출고": {"file": "생산출고_template.xlsx", "start_row": 4},
                        "상용직정보": {"file": "상용직정보_template.xlsx", "start_row": 8},
                        "품목등록": {"file": "품목등록_template.xlsx", "start_row": 4},
                        "BOM등록": {"file": "BOM등록_template.xlsx", "start_row": 4},
                        "품목군등록": {"file": "품목군등록_template.xlsx", "start_row": 4},
                        "기초재고": {"file": "기초재고_template.xlsx", "start_row": 4},
                        "관리내역등록": {"file": "관리내역등록_template.xlsx", "start_row": 4},
                        "물류담당자등록": {"file": "물류담당자등록_template.xlsx", "start_row": 4},
                        "고객별물류담당자등록": {"file": "고객별물류담당자등록_template.xlsx", "start_row": 4},
                        "프로젝트등록": {"file": "프로젝트등록_template.xlsx", "start_row": 4},
                        "창고": {"file": "창고_template.xlsx", "start_row": 4},
                        "공정": {"file": "공정_template.xlsx", "start_row": 4},
                        "발주등록": {"file": "발주등록_template.xlsx", "start_row": 4},
                        "수금등록": {"file": "수금등록_template.xlsx", "start_row": 4},
                        "입고처리": {"file": "입고처리_template.xlsx", "start_row": 4},
                        "출고처리": {"file": "출고처리_template.xlsx", "start_row": 4},
                        "재고조정": {"file": "재고조정_template.xlsx", "start_row": 4},
                        "재고이동": {"file": "재고이동_template.xlsx", "start_row": 4},
                        "회계초기이월": {"file": "회계초기이월_template.xlsx", "start_row": 4},
                        "자동전표처리": {"file": "자동전표처리_template.xlsx", "start_row": 4},
                        "납품처등록": {"file": "납품처등록_template.xlsx", "start_row": 4},
                    }

                    config = template_config.get(query_name)
                    if not config:
                        app.logger.error(f"템플릿 설정이 없음: {query_name}")
                        return jsonify({"error": f"'{query_name}'에 대한 엑셀 템플릿 설정이 없습니다."}), 400

                    template_filename = config.get("file")
                    start_row = config.get("start_row", 4)
                    template_path = os.path.join(TEMPLATE_FOLDER, template_filename)

                    if not os.path.exists(template_path):
                        app.logger.error(f"템플릿 파일을 찾을 수 없음: {template_path}")
                        return jsonify({"error": f"엑셀 템플릿 파일({template_filename})을 찾을 수 없습니다."}), 404

                    workbook = openpyxl.load_workbook(template_path)
                    worksheet = workbook.active

                    for r_idx, row_data in enumerate(chunk_df.itertuples(index=False), start=start_row):
                        for c_idx, cell_value in enumerate(row_data, 1):
                            if pd.isna(cell_value):
                                cell_value = ""
                            worksheet.cell(row=r_idx, column=c_idx, value=cell_value)

                    excel_buffer = io.BytesIO()
                    workbook.save(excel_buffer)
                    excel_buffer.seek(0)

                    zf.writestr(file_name, excel_buffer.getvalue())

        zip_buffer.seek(0)
        return send_file(zip_buffer, as_attachment=True, download_name=download_filename, mimetype='application/zip')

    except Exception as e:
        app.logger.error(f"엑셀 생성 중 오류: {e}", exc_info=True)
        return jsonify({"error": f"엑셀 생성 오류: {str(e)}"}), 500
    finally:
        if filepath and os.path.exists(filepath):
            os.remove(filepath)

# --- Flask 서버 실행 ---
if __name__ == '__main__':

    app.run(host='0.0.0.0', port=5000, debug=True)



